from pypdf import PdfReader
from openpyxl import load_workbook
import csv
import zipfile


def test_csv():
    with zipfile.ZipFile('content_folder/archive_folder/archive.zip') as zip_file:
        with zip_file.open('csv') as csv_file:
            content = csv_file.read().decode('utf-8-sig')
            csvreader = list(csv.reader(content.splitlines()))
            second_row = csvreader[1]
            assert 'OU001' in second_row[0]

def test_pdf():
    with (zipfile.ZipFile('content_folder/archive_folder/archive.zip') as zip_file):
        with zip_file.open('pdf') as pdf_file:
            reader = PdfReader(pdf_file)
            text = reader.pages[0].extract_text()
            assert "Commercial Invoice" in text


def test_xlsx():
    with zipfile.ZipFile('content_folder/archive_folder/archive.zip') as zip_file:
        with zip_file.open('xlsx') as xlsx_file:
            workbook = load_workbook(xlsx_file)
            sheet = workbook.active
            assert 'OU001' in sheet.cell(row=2, column=1).value
            assert 'Коммерческий департамент' in sheet.cell(row=2, column=3).value
